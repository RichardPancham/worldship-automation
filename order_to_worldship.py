"""
UPS WorldShip Auto-Import Script
----------------------------------
How it works:
  1. Customer places an order (web form, phone, internal system, etc.)
  2. This script writes the order to a CSV file in WorldShip's import folder
  3. WorldShip's built-in Keyed Import detects the file and auto-populates
     all shipping fields — no manual typing needed
  4. Staff just hit Process/Print and the label comes out

WorldShip CSV import folder (default location):
  C:\\ProgramData\\UPS\\WSTD\\ImpExp\\AcctPkgs\\Sample Order Import 1\\worldship.csv

Column names match WorldShip's official import map format.
"""

import csv
import os
import openpyxl
from datetime import datetime

# ---------------------------------------------------------------
# CONFIG — update these paths if needed
# ---------------------------------------------------------------
WORLDSHIP_CSV = r"C:\ProgramData\UPS\WSTD\ImpExp\AcctPkgs\Sample Order Import 1\worldship.csv"
ORDERS_EXCEL  = "orders.xlsx"
LOG_FILE      = "import_log.txt"

# WorldShip expects these exact column headers for its default import map
WORLDSHIP_COLUMNS = [
    "OrderNumber",
    "CompanyOrName",
    "Address1",
    "Address2",
    "City",
    "StateProvinceCounty",
    "PostalCode",
    "CountryTerritory",
    "Telephone",
    "ResidentialIndicator",
    "ServiceType",
    "PackageType",
    "Weight",
    "ReferenceNumber1",
]

# ---------------------------------------------------------------
# HELPERS
# ---------------------------------------------------------------
def log(msg):
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{ts}] {msg}"
    print(line)
    with open(LOG_FILE, "a") as f:
        f.write(line + "\n")


def get_pending_orders():
    """Read orders.xlsx and return all rows with Status = Pending"""
    wb = openpyxl.load_workbook(ORDERS_EXCEL)
    ws = wb.active
    pending = []

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0]:
            continue
        order_id, name, address, city, province, postal, country, phone, weight, service, status = row[:11]
        if str(status).strip().lower() == "pending":
            pending.append({
                "row":      i,
                "order_id": str(order_id),
                "name":     str(name),
                "address":  str(address),
                "city":     str(city),
                "province": str(province),
                "postal":   str(postal),
                "country":  str(country) if country else "CA",
                "phone":    str(phone) if phone else "",
                "weight":   str(weight),
                "service":  str(service),
            })

    wb.close()
    return pending


def write_worldship_csv(orders):
    """
    Write pending orders to WorldShip's import CSV.
    WorldShip will pick these up automatically on next Keyed Import.
    """
    os.makedirs(os.path.dirname(WORLDSHIP_CSV), exist_ok=True)

    # Check if file exists to decide whether to write header
    file_exists = os.path.exists(WORLDSHIP_CSV)

    with open(WORLDSHIP_CSV, "a", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=WORLDSHIP_COLUMNS)

        if not file_exists:
            writer.writeheader()
            log("Created new WorldShip import CSV with headers")

        for order in orders:
            writer.writerow({
                "OrderNumber":           order["order_id"],
                "CompanyOrName":         order["name"],
                "Address1":              order["address"],
                "Address2":              "",
                "City":                  order["city"],
                "StateProvinceCounty":   order["province"],
                "PostalCode":            order["postal"],
                "CountryTerritory":      order["country"],
                "Telephone":             order["phone"],
                "ResidentialIndicator":  "Y",
                "ServiceType":           map_service(order["service"]),
                "PackageType":           "CP",   # Customer Packaging
                "Weight":                order["weight"],
                "ReferenceNumber1":      order["order_id"],
            })
            log(f"Written to WorldShip CSV: {order['order_id']} — {order['name']}, {order['city']}")


def map_service(service_name):
    """
    Map human-readable service names to WorldShip service codes.
    These are the official UPS service type codes.
    """
    mapping = {
        "ups ground":         "GND",
        "ups 2nd day air":    "2DA",
        "ups next day air":   "1DA",
        "ups 3 day select":   "3DS",
        "ups worldwide":      "WXS",
    }
    return mapping.get(service_name.lower(), "GND")  # default to Ground


def mark_exported(row_number, order_id):
    """Update Status in orders.xlsx to 'Exported to WorldShip'"""
    wb = openpyxl.load_workbook(ORDERS_EXCEL)
    ws = wb.active
    ws.cell(row=row_number, column=11).value = "Exported to WorldShip"
    ws.cell(row=row_number, column=12).value = datetime.now().strftime("%Y-%m-%d %H:%M")
    wb.save(ORDERS_EXCEL)
    wb.close()
    log(f"Marked {order_id} as Exported in orders.xlsx")


# ---------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------
def run():
    log("=== WorldShip Import Script Started ===")
    log(f"Reading orders from: {ORDERS_EXCEL}")
    log(f"Writing to WorldShip CSV: {WORLDSHIP_CSV}")

    pending = get_pending_orders()

    if not pending:
        log("No pending orders found. Nothing to export.")
        return

    log(f"Found {len(pending)} pending order(s): {[o['order_id'] for o in pending]}")

    write_worldship_csv(pending)

    for order in pending:
        mark_exported(order["row"], order["order_id"])

    log(f"Done — {len(pending)} order(s) exported to WorldShip.")
    log("In WorldShip: go to Import/Export Data > Keyed Import > enter order number to load.")
    print("\n" + "="*55)
    print(f"  {len(pending)} order(s) ready in WorldShip import file.")
    print("  Open WorldShip > Import/Export > Keyed Import")
    print("  Type the order number — fields auto-populate!")
    print("="*55)


if __name__ == "__main__":
    run()
