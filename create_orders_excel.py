"""
Run this once to create the orders.xlsx template.
Columns match what order_to_worldship.py expects to read.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Orders"

headers = [
    "Order ID", "Name", "Address", "City", "Province",
    "Postal Code", "Country", "Phone", "Weight (lbs)",
    "Service", "Status", "Exported At"
]
col_widths = [12, 20, 25, 15, 12, 14, 10, 16, 14, 20, 22, 18]

header_fill = PatternFill("solid", fgColor="1D9E75")
header_font = Font(bold=True, color="FFFFFF", size=11)

for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")
    ws.column_dimensions[cell.column_letter].width = w

sample_data = [
    ["#1001", "Mike Johnson",  "123 Kennedy Rd",    "Brampton",    "ON", "L6T 2H5", "CA", "4161234567", 3.1, "UPS Ground",       "Exported to WorldShip", "2024-01-15 09:12"],
    ["#1002", "Sara Lee",      "456 Hurontario St", "Mississauga", "ON", "L5B 3K2", "CA", "9059876543", 7.8, "UPS 2nd Day Air",  "Exported to WorldShip", "2024-01-15 09:45"],
    ["#1003", "David Nguyen",  "789 Yonge St",      "Toronto",     "ON", "M4W 2G8", "CA", "4162223333", 2.5, "UPS Ground",       "Pending",               ""],
    ["#1004", "Priya Sharma",  "22 Eglinton Ave W", "Toronto",     "ON", "M4R 1K8", "CA", "4165554444", 5.0, "UPS Next Day Air", "Pending",               ""],
]

for row in sample_data:
    ws.append(row)

ws.freeze_panes = "A2"
wb.save("orders.xlsx")
print("orders.xlsx created — 2 sample Pending orders ready to export.")
